from flask import render_template, request, send_file, flash, jsonify
from flask_login import login_required
from datetime import datetime
import pandas as pd
import io
import logging
from openpyxl.styles import Font, PatternFill

# Импортируем функцию подключения и параметры из глобального модуля redmine
from redmine import (
    get_connection,
    db_redmine_host,
    db_redmine_port,
    db_redmine_user_name,
    db_redmine_password,
    db_redmine_name
)

from . import reports_bp
logger = logging.getLogger(__name__)

def get_db_connection():
    return get_connection(
        db_redmine_host,
        db_redmine_user_name,
        db_redmine_password,
        db_redmine_name,
        port=db_redmine_port
    )

def get_projects():
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, name FROM redmine.projects WHERE status = 1 ORDER BY name")
            return cursor.fetchall()
    finally:
        conn.close()

def get_statuses():
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, name, is_closed FROM redmine.issue_statuses ORDER BY position")
            return cursor.fetchall()
    finally:
        conn.close()

def get_executors():
    conn = get_db_connection()
    if not conn:
        return []
    try:
        with conn.cursor() as cursor:
            # Выбираем активных пользователей, которые есть в задачах (упрощенно) или просто всех активных
            # Для производительности выберем всех активных пользователей, у которых есть lastname
            cursor.execute("SELECT id, CONCAT(lastname, ' ', firstname) as name FROM redmine.users WHERE status = 1 AND type = 'User' ORDER BY lastname, firstname")
            return cursor.fetchall()
    finally:
        conn.close()

@reports_bp.route("/redmine-report", methods=["GET"])
@login_required
def redmine_report():
    projects = get_projects()
    statuses = get_statuses()
    executors = get_executors()

    # Параметры фильтра
    project_id = request.args.get('project_id', type=int)
    period_type = request.args.get('period_type', 'created') # created or closed
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    # Обработка фильтра статуса (быстрые фильтры + конкретные статусы)
    status_filter = request.args.get('status_filter', '')
    status_id = None
    is_closed_filter = None
    if status_filter:
        if status_filter == 'is_open':
            is_closed_filter = 0
        elif status_filter == 'is_closed':
            is_closed_filter = 1
        else:
            try:
                status_id = int(status_filter)
            except ValueError:
                status_id = None

    # Новые параметры
    report_year = request.args.get('report_year', type=int) or None
    months = request.args.getlist('months') # Список строк "1", "2"...
    executor_id = request.args.get('executor_id', type=int)
    group_mode = request.args.get('group_mode', 'EXECUTOR')

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int) # Уменьшил дефолт т.к. группы большие

    # Логика дефолтов - если год не выбран (пустая строка), используем None (все время)
    # months по умолчанию - все месяцы
    current_year = datetime.now().year
    if not months:
        months = [str(i) for i in range(1, 13)]

    report_data = [] # Список групп (верхнего уровня)
    total_rows = 0 # Количество групп верхнего уровня
    total_issues_count = 0 # Общий итог

    if project_id:
        conn = get_db_connection()
        if conn:
            try:
                period_field = 'i.created_on' if period_type == 'created' else 'i.closed_on'

                # Общие условия фильтрации (WHERE)
                # Выносим в отдельную строку для переиспользования
                where_clause = """
                    WHERE i.project_id = %s
                """
                params = [project_id]

                # Фильтр по году (если указан) и месяцам
                if report_year:
                    where_clause += f" AND YEAR({period_field}) = %s"
                    params.append(report_year)

                if months:
                    placeholders = ','.join(['%s'] * len(months))
                    where_clause += f" AND MONTH({period_field}) IN ({placeholders})"
                    params.extend(months)

                # Дополнительный фильтр date_from/date_to (пересечение)
                if date_from:
                    where_clause += f" AND {period_field} >= %s"
                    params.append(date_from + " 00:00:00")
                if date_to:
                    where_clause += f" AND {period_field} < DATE_ADD(%s, INTERVAL 1 DAY)"
                    params.append(date_to + " 00:00:00")

                if status_id:
                    where_clause += " AND i.status_id = %s"
                    params.append(status_id)
                if is_closed_filter is not None:
                    where_clause += " AND s.is_closed = %s"
                    params.append(is_closed_filter)
                if period_type == 'closed':
                    where_clause += " AND i.closed_on IS NOT NULL AND s.is_closed = 1"

                if executor_id:
                    where_clause += " AND COALESCE(i.assigned_to_id, i.easy_closed_by_id) = %s"
                    params.append(executor_id)

                # Шаг 1: Считаем общий итог по всем данным
                total_sql = f"""
                    SELECT COUNT(*) as total
                    FROM redmine.issues i
                    LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
                    {where_clause}
                """
                with conn.cursor() as cursor:
                    cursor.execute(total_sql, params)
                    res = cursor.fetchone()
                    total_issues_count = res['total'] if res else 0

                # Определение ключей группировки в зависимости от режима
                # group_mode: EXECUTOR, MONTH, EXECUTOR_MONTH, MONTH_EXECUTOR

                key_expression = ""
                name_expression = ""
                group_by_clause = ""
                order_by_clause = ""

                second_level_needed = False

                if group_mode == 'EXECUTOR' or group_mode == 'EXECUTOR_MONTH':
                    # Use 0 as key for NULL assignees or empty names to ensure single group
                    key_expression = "CASE WHEN COALESCE(NULLIF(CONCAT_WS(' ', u.lastname, u.firstname), ''), 'Исполнитель не назначен') = 'Исполнитель не назначен' THEN 0 ELSE COALESCE(i.assigned_to_id, i.easy_closed_by_id) END"
                    name_expression = "COALESCE(NULLIF(CONCAT_WS(' ', u.lastname, u.firstname), ''), 'Исполнитель не назначен')"

                    top_level_select = f"{key_expression} as key_id, {name_expression} as key_name"
                    group_by_clause = f"{key_expression}, {name_expression}"
                    order_by_clause = "key_name"

                    if group_mode == 'EXECUTOR_MONTH':
                        second_level_needed = True

                elif group_mode == 'MONTH' or group_mode == 'MONTH_EXECUTOR':
                    key_expression = f"DATE_FORMAT({period_field}, '%%Y-%%m')"
                    name_expression = key_expression

                    top_level_select = f"{key_expression} as key_name, {key_expression} as key_id"
                    group_by_clause = key_expression
                    order_by_clause = key_expression

                    if group_mode == 'MONTH_EXECUTOR':
                        second_level_needed = True

                else:
                    raise ValueError(f"Unknown group_mode: {group_mode}")

                # Шаг 2: Пагинация верхнего уровня
                # Используем прямое выражение (key_expression) вместо алиаса
                count_groups_sql = f"""
                    SELECT COUNT(DISTINCT {key_expression}) as count
                    FROM redmine.issues i
                    LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
                    LEFT JOIN redmine.users u ON u.id = COALESCE(i.assigned_to_id, i.easy_closed_by_id)
                    {where_clause}
                """
                with conn.cursor() as cursor:
                    cursor.execute(count_groups_sql, params)
                    res = cursor.fetchone()
                    total_rows = res['count'] if res else 0

                # Получаем ключи для текущей страницы
                page_keys_sql = f"""
                    SELECT {top_level_select}, COUNT(*) as total_cnt
                    FROM redmine.issues i
                    LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
                    LEFT JOIN redmine.users u ON u.id = COALESCE(i.assigned_to_id, i.easy_closed_by_id)
                    {where_clause}
                    GROUP BY {group_by_clause}
                    ORDER BY {order_by_clause}
                    LIMIT %s OFFSET %s
                """
                page_params = params.copy()
                page_params.extend([page_size, (page - 1) * page_size])

                top_level_groups = []
                with conn.cursor() as cursor:
                    cursor.execute(page_keys_sql, page_params)
                    top_level_groups = cursor.fetchall()

                # Шаг 3: Собираем данные (второй уровень, если нужен)
                # Для каждой группы верхнего уровня формируем структуру
                for group in top_level_groups:
                    # Translate 0 to unassigned for frontend/API compatibility
                    group_key = group['key_id']
                    if group_key == 0 or group_key == '0':
                        group_key = 'unassigned'
                    group_item = {
                        'id': group_key,
                        'name': group['key_name'],
                        'count': group['total_cnt'],
                        'children': []
                    }

                    if second_level_needed:
                        # Запрос для второго уровня
                        # Нужно добавить условие фильтрации по родителю
                        sub_where = where_clause
                        sub_params = params.copy()

                        sub_select = ""
                        sub_group = ""
                        sub_order = ""

                        if group_mode == 'EXECUTOR_MONTH':
                            # Родитель - Executor, Дети - Month
                            if group_item['id'] == 'unassigned':
                                sub_where += " AND COALESCE(i.assigned_to_id, i.easy_closed_by_id) IS NULL"
                            else:
                                sub_where += " AND COALESCE(i.assigned_to_id, i.easy_closed_by_id) = %s"
                                sub_params.append(group_item['id'])

                            sub_select = f"DATE_FORMAT({period_field}, '%%Y-%%m') as key_name, DATE_FORMAT({period_field}, '%%Y-%%m') as key_id"
                            sub_group = "key_id"
                            sub_order = "key_id"

                        elif group_mode == 'MONTH_EXECUTOR':
                            # Родитель - Month, Дети - Executor
                            # Filter by month str check?
                            # period_field is date. DATE_FORMAT returns string.
                            # Better use DATE_FORMAT in WHERE or YEAR/MONTH
                            # group_item['id'] is 'YYYY-MM'
                            y, m = group_item['id'].split('-')
                            sub_where += f" AND YEAR({period_field}) = %s AND MONTH({period_field}) = %s"
                            sub_params.extend([y, m])

                            sub_select = "CASE WHEN COALESCE(NULLIF(CONCAT_WS(' ', u.lastname, u.firstname), ''), 'Исполнитель не назначен') = 'Исполнитель не назначен' THEN 0 ELSE COALESCE(i.assigned_to_id, i.easy_closed_by_id) END as key_id, COALESCE(NULLIF(CONCAT_WS(' ', u.lastname, u.firstname), ''), 'Исполнитель не назначен') as key_name"
                            sub_group = "key_id, key_name"
                            sub_order = "key_name"

                        children_sql = f"""
                            SELECT {sub_select}, COUNT(*) as cnt
                            FROM redmine.issues i
                            LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
                            LEFT JOIN redmine.users u ON u.id = COALESCE(i.assigned_to_id, i.easy_closed_by_id)
                            {sub_where}
                            GROUP BY {sub_group}
                            ORDER BY {sub_order}
                        """
                        with conn.cursor() as cursor:
                            cursor.execute(children_sql, sub_params)
                            children = cursor.fetchall()
                            for child in children:
                                child_key = child['key_id']
                                if child_key == 0 or child_key == '0':
                                    child_key = 'unassigned'
                                group_item['children'].append({
                                    'id': child_key,
                                    'name': child['key_name'],
                                    'count': child['cnt']
                                })

                    report_data.append(group_item)

                # Merge duplicate groups (safety net)
                merged_data = {}
                for g in report_data:
                    gid = g['id']
                    if gid in merged_data:
                        merged_data[gid]['count'] += g['count']
                        merged_data[gid]['children'].extend(g['children'])
                    else:
                        merged_data[gid] = g
                report_data = list(merged_data.values())

            except Exception as e:
                logger.error("Error in redmine_report: %s", e, exc_info=True)
                flash(f"Ошибка формирования отчета: {e}", "danger")
            finally:
                conn.close()

    return render_template(
        'reports/redmine_report.html',
        projects=projects,
        statuses=statuses,
        executors=executors,
        report_data=report_data,
        total_rows=total_rows,
        total_issues_count=total_issues_count,
        # Params
        project_id=project_id,
        period_type=period_type,
        date_from=date_from,
        date_to=date_to,
        status_id=status_id,
        status_filter=status_filter,
        report_year=report_year,
        months=months,
        executor_id=executor_id,
        group_mode=group_mode,
        page=page,
        page_size=page_size,
        total_pages=(total_rows + page_size - 1) // page_size if page_size else 0,
        count_notifications=0,
        current_year=current_year
    )

@reports_bp.route("/redmine-report/details", methods=["GET"])
def redmine_report_details():
    """API для получения списка задач (Drill-down)"""
    try:
        project_id = request.args.get('project_id', type=int)
        period_type = request.args.get('period_type', 'created')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')

        # Обработка фильтра статуса
        status_filter = request.args.get('status_filter', '')
        status_id = None
        is_closed_filter = None
        if status_filter:
            if status_filter == 'is_open':
                is_closed_filter = 0
            elif status_filter == 'is_closed':
                is_closed_filter = 1
            else:
                try:
                    status_id = int(status_filter)
                except ValueError:
                    status_id = None

        report_year = request.args.get('report_year', type=int) or None
        months = request.args.getlist('months[]') # В JS массивы передаются с []
        if not months: months = request.args.getlist('months') # Fallback

        executor_id_filter = request.args.get('executor_id', type=int) # Глобальный фильтр

        # Контекстные фильтры (от группы)
        ctx_executor_id = request.args.get('ctx_executor_id')  # Убрали type=int для поддержки 'unassigned'
        ctx_ym = request.args.get('ctx_ym') # YYYY-MM

        detail_page = request.args.get('detail_page', 1, type=int)
        detail_page_size = request.args.get('detail_page_size', 20, type=int)

        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'db error'}), 500

        period_field = 'i.created_on' if period_type == 'created' else 'i.closed_on'

        where_clause = """
            WHERE i.project_id = %s
        """
        params = [project_id]

        # Основные фильтры
        if report_year:
            where_clause += f" AND YEAR({period_field}) = %s"
            params.append(report_year)

        # Months logic: filter by list AND ctx_ym intersection?
        # Typically ctx_ym dictates the specific month.
        # If ctx_ym is present, it overrides the list for this specific query scope basically.

        if ctx_ym:
            y, m = ctx_ym.split('-')
            where_clause += f" AND YEAR({period_field}) = %s AND MONTH({period_field}) = %s"
            params.extend([y, m])
        elif months:
            placeholders = ','.join(['%s'] * len(months))
            where_clause += f" AND MONTH({period_field}) IN ({placeholders})"
            params.extend(months)

        if date_from:
            where_clause += f" AND {period_field} >= %s"
            params.append(date_from + " 00:00:00")
        if date_to:
            where_clause += f" AND {period_field} < DATE_ADD(%s, INTERVAL 1 DAY)"
            params.append(date_to + " 00:00:00")
        if status_id:
            where_clause += " AND i.status_id = %s"
            params.append(status_id)
        if is_closed_filter is not None:
            where_clause += " AND s.is_closed = %s"
            params.append(is_closed_filter)
        if period_type == 'closed':
            where_clause += " AND i.closed_on IS NOT NULL AND s.is_closed = 1"

        # Executor filters
        target_executor_id = ctx_executor_id if ctx_executor_id else executor_id_filter
        if target_executor_id == 'unassigned':
            # Special case: filter for unassigned issues (assigned_to_id IS NULL)
            where_clause += " AND COALESCE(i.assigned_to_id, i.easy_closed_by_id) IS NULL"
        elif target_executor_id:
            where_clause += " AND COALESCE(i.assigned_to_id, i.easy_closed_by_id) = %s"
            params.append(target_executor_id)

        # Count total
        count_sql = f"""
            SELECT COUNT(*) as total
            FROM redmine.issues i
            LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
            {where_clause}
        """

        total_issues = 0
        with conn.cursor() as cursor:
            cursor.execute(count_sql, params)
            res = cursor.fetchone()
            total_issues = res['total'] if res else 0

        # Select data
        # Fields: ID, Subject, Status, date, closed, executor
        sql = f"""
            SELECT
                i.id,
                i.subject,
                s.name as status_name,
                DATE_FORMAT(i.created_on, '%%Y-%%m-%%d %%H:%%i') as created_on,
                DATE_FORMAT(i.closed_on, '%%Y-%%m-%%d %%H:%%i') as closed_on,
                CONCAT_WS(' ', u.lastname, u.firstname) as executor_name
            FROM redmine.issues i
            LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
            LEFT JOIN redmine.users u ON u.id = COALESCE(i.assigned_to_id, i.easy_closed_by_id)
            {where_clause}
            ORDER BY {period_field} DESC
            LIMIT %s OFFSET %s
        """
        params.extend([detail_page_size, (detail_page - 1) * detail_page_size])

        issues = []
        with conn.cursor() as cursor:
            cursor.execute(sql, params)
            issues = cursor.fetchall()

        return jsonify({
            'issues': issues,
            'total_issues': total_issues,
            'page': detail_page,
            'page_size': detail_page_size,
            'total_pages': (total_issues + detail_page_size - 1) // detail_page_size if detail_page_size else 0
        })

    except Exception as e:
        logger.error("Details API Error: %s", e, exc_info=True)
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn:
            conn.close()

@reports_bp.route("/redmine-report/export", methods=["GET"])
def redmine_report_export():
    try:
        project_id = request.args.get('project_id', type=int)
        if not project_id:
            return "Project required", 400

        # Params same as main report
        period_type = request.args.get('period_type', 'created')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')

        # Обработка фильтра статуса
        status_filter = request.args.get('status_filter', '')
        status_id = None
        is_closed_filter = None
        if status_filter:
            if status_filter == 'is_open':
                is_closed_filter = 0
            elif status_filter == 'is_closed':
                is_closed_filter = 1
            else:
                try:
                    status_id = int(status_filter)
                except ValueError:
                    status_id = None

        report_year = request.args.get('report_year', type=int) or None
        months = request.args.getlist('months')
        if not months: months = [str(i) for i in range(1, 13)]
        executor_id = request.args.get('executor_id', type=int)
        group_mode = request.args.get('group_mode', 'EXECUTOR')

        conn = get_db_connection()
        if not conn: return "DB Error", 500

        period_field = 'i.created_on' if period_type == 'created' else 'i.closed_on'

        # Construct WHERE
        where_params = [project_id]
        where_clause = """
            WHERE i.project_id = %s
        """
        if report_year:
            where_clause += f" AND YEAR({period_field}) = %s"
            where_params.append(report_year)

        if months:
            placeholders = ','.join(['%s'] * len(months))
            where_clause += f" AND MONTH({period_field}) IN ({placeholders})"
            where_params.extend(months)

        if date_from:
            where_clause += f" AND {period_field} >= %s"
            where_params.append(date_from + " 00:00:00")
        if date_to:
            where_clause += f" AND {period_field} < DATE_ADD(%s, INTERVAL 1 DAY)"
            where_params.append(date_to + " 00:00:00")
        if status_id:
            where_clause += " AND i.status_id = %s"
            where_params.append(status_id)
        if is_closed_filter is not None:
            where_clause += " AND s.is_closed = %s"
            where_params.append(is_closed_filter)
        if period_type == 'closed':
            where_clause += " AND i.closed_on IS NOT NULL AND s.is_closed = 1"
        if executor_id:
            where_clause += " AND COALESCE(i.assigned_to_id, i.easy_closed_by_id) = %s"
            where_params.append(executor_id)

        # 1. Aggregates for "Report" Sheet
        # Just dump flat and simple? Or try to replicate hierarchy?
        # User said "table aggregation in selected group_mode".
        # Simplified: Dump flat grouping based on mode components.

        group_sql = ""
        agg_params = where_params.copy()

        if group_mode == 'EXECUTOR':
            group_sql = f"""
                SELECT COALESCE(NULLIF(CONCAT_WS(' ', u.lastname, u.firstname), ''), 'Исполнитель не назначен') as col1, NULL as col2, COUNT(*) as cnt
                FROM redmine.issues i
                LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
                LEFT JOIN redmine.users u ON u.id = COALESCE(i.assigned_to_id, i.easy_closed_by_id)
                {where_clause}
                GROUP BY col1
                ORDER BY col1
            """
            headers = ["Исполнитель", "", "Кол-во"]
        elif group_mode == 'MONTH':
            group_sql = f"""
                SELECT DATE_FORMAT({period_field}, '%%Y-%%m') as col1, NULL as col2, COUNT(*) as cnt
                FROM redmine.issues i
                LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
                LEFT JOIN redmine.users u ON u.id = COALESCE(i.assigned_to_id, i.easy_closed_by_id)
                {where_clause}
                GROUP BY col1
                ORDER BY col1
            """
            headers = ["Месяц", "", "Кол-во"]
        elif group_mode == 'EXECUTOR_MONTH':
            group_sql = f"""
                SELECT COALESCE(NULLIF(CONCAT_WS(' ', u.lastname, u.firstname), ''), 'Исполнитель не назначен') as col1, DATE_FORMAT({period_field}, '%%Y-%%m') as col2, COUNT(*) as cnt
                FROM redmine.issues i
                LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
                LEFT JOIN redmine.users u ON u.id = COALESCE(i.assigned_to_id, i.easy_closed_by_id)
                {where_clause}
                GROUP BY col1, col2
                ORDER BY col1, col2
            """
            headers = ["Исполнитель", "Месяц", "Кол-во"]
        else: # MONTH_EXECUTOR
            group_sql = f"""
                SELECT DATE_FORMAT({period_field}, '%%Y-%%m') as col1, COALESCE(NULLIF(CONCAT_WS(' ', u.lastname, u.firstname), ''), 'Исполнитель не назначен') as col2, COUNT(*) as cnt
                FROM redmine.issues i
                LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
                LEFT JOIN redmine.users u ON u.id = COALESCE(i.assigned_to_id, i.easy_closed_by_id)
                {where_clause}
                GROUP BY col1, col2
                ORDER BY col1, col2
            """
            headers = ["Месяц", "Исполнитель", "Кол-во"]

        with conn.cursor() as cursor:
            cursor.execute(group_sql, agg_params)
            rows_agg = cursor.fetchall()

        # 2. Details for "Issues" Sheet
        issues_sql = f"""
            SELECT
                i.id,
                i.subject,
                s.name as status_name,
                i.created_on,
                i.closed_on,
                COALESCE(NULLIF(CONCAT_WS(' ', u.lastname, u.firstname), ''), 'Исполнитель не назначен') as executor_name
            FROM redmine.issues i
            LEFT JOIN redmine.issue_statuses s ON s.id = i.status_id
            LEFT JOIN redmine.users u ON u.id = COALESCE(i.assigned_to_id, i.easy_closed_by_id)
            {where_clause}
            ORDER BY {period_field} DESC
        """
        with conn.cursor() as cursor:
            cursor.execute(issues_sql, where_params)
            rows_issues = cursor.fetchall()

        # CREATE EXCEL
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Report
            worksheet = writer.book.create_sheet('Report')
            writer.sheets['Report'] = worksheet

            # Params
            start_row = 1
            worksheet.cell(row=start_row, column=1, value="Параметры отчёта").font = Font(bold=True, size=14)
            start_row += 1
            params_display = {
               "Project ID": project_id,
               "Period Type": period_type,
               "Year": report_year,
               "Months": ",".join(months) if months else "All",
               "Date From": date_from,
               "Date To": date_to,
               "Status": status_filter or "All",
               "Executor": executor_id or "All",
               "Generated": datetime.now().strftime("%Y-%m-%d %H:%M")
            }
            for k, v in params_display.items():
                worksheet.cell(row=start_row, column=1, value=k)
                worksheet.cell(row=start_row, column=2, value=str(v))
                start_row += 1
            start_row += 1

            # Table Header
            for c, h in enumerate(headers, 1):
                cell = worksheet.cell(row=start_row, column=c, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            start_row += 1

            # Table Data
            for r in rows_agg:
                worksheet.cell(row=start_row, column=1, value=r['col1'])
                worksheet.cell(row=start_row, column=2, value=r['col2'])
                worksheet.cell(row=start_row, column=3, value=r['cnt'])
                start_row += 1

            # Sheet 2: Issues
            df_issues = pd.DataFrame(rows_issues)
            if not df_issues.empty:
                # Rename columns for display
                df_issues.columns = ['ID', 'Тема', 'Статус', 'Создана', 'Закрыта', 'Исполнитель']
                # Convert timestamps to string to avoid tz issues in excel sometimes
                df_issues['Создана'] = df_issues['Создана'].astype(str)
                df_issues['Закрыта'] = df_issues['Закрыта'].astype(str)
                df_issues.to_excel(writer, sheet_name='Issues', index=False)
            else:
                writer.book.create_sheet('Issues')
                writer.sheets['Issues'].cell(row=1, column=1, value="Нет задач")

        output.seek(0)
        filename = f"redmine_report_v2_{project_id}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        logger.error("Export Error: %s", e, exc_info=True)
        return str(e), 500
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@reports_bp.route("/redmine-report/stale-tasks", methods=["GET"])
def redmine_stale_tasks():
    """API для получения списка зависших задач"""
    try:
        project_id = request.args.get('project_id', type=int)
        if not project_id:
            return jsonify({'error': 'project_id required'}), 400

        stale_days = request.args.get('stale_days', 7, type=int)
        status_id = request.args.get('status_id', type=int)
        executor_id = request.args.get('executor_id', type=int)
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 50, type=int)

        conn = get_db_connection()
        if not conn:
            return jsonify({'error': 'db error'}), 500

        where_clause = "WHERE i.project_id = %s AND s.is_closed = 0"
        params = [project_id, stale_days]

        where_clause += " AND COALESCE(DATEDIFF(NOW(), i.easy_status_updated_on), DATEDIFF(NOW(), i.updated_on)) >= %s"

        if status_id:
            where_clause += " AND i.status_id = %s"
            params.append(status_id)
        if executor_id:
            where_clause += " AND i.assigned_to_id = %s"
            params.append(executor_id)

        count_sql = f"""
            SELECT COUNT(*) as total
            FROM redmine.issues i
            JOIN redmine.issue_statuses s ON s.id = i.status_id
            {where_clause}
        """
        with conn.cursor() as cursor:
            cursor.execute(count_sql, params)
            total = cursor.fetchone()['total']

        sql = f"""
            SELECT
                i.id,
                i.subject,
                s.name as status_name,
                s.id as status_id,
                DATE_FORMAT(i.created_on, '%%Y-%%m-%%d %%H:%%i') as created_on,
                DATEDIFF(NOW(), i.created_on) as age_days,
                COALESCE(DATEDIFF(NOW(), i.easy_status_updated_on), DATEDIFF(NOW(), i.updated_on)) as stale_days,
                CONCAT_WS(' ', u.lastname, u.firstname) as assigned_to_name
            FROM redmine.issues i
            JOIN redmine.issue_statuses s ON s.id = i.status_id
            LEFT JOIN redmine.users u ON u.id = i.assigned_to_id
            {where_clause}
            ORDER BY stale_days DESC
            LIMIT %s OFFSET %s
        """
        params.extend([page_size, (page - 1) * page_size])

        with conn.cursor() as cursor:
            cursor.execute(sql, params)
            issues = cursor.fetchall()

        return jsonify({
            'issues': issues,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size if page_size else 0
        })

    except Exception as e:
        logger.error("Stale tasks API Error: %s", e, exc_info=True)
        return jsonify({'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@reports_bp.route("/redmine-report/stale-tasks/export", methods=["GET"])
def redmine_stale_tasks_export():
    """Экспорт зависших задач в Excel"""
    try:
        project_id = request.args.get('project_id', type=int)
        if not project_id:
            return "Project required", 400

        stale_days = request.args.get('stale_days', 7, type=int)
        status_id = request.args.get('status_id', type=int)
        executor_id = request.args.get('executor_id', type=int)

        conn = get_db_connection()
        if not conn:
            return "DB Error", 500

        where_clause = "WHERE i.project_id = %s AND s.is_closed = 0"
        params = [project_id, stale_days]

        where_clause += " AND COALESCE(DATEDIFF(NOW(), i.easy_status_updated_on), DATEDIFF(NOW(), i.updated_on)) >= %s"

        if status_id:
            where_clause += " AND i.status_id = %s"
            params.append(status_id)
        if executor_id:
            where_clause += " AND i.assigned_to_id = %s"
            params.append(executor_id)

        sql = f"""
            SELECT
                i.id,
                i.subject,
                s.name as status_name,
                DATEDIFF(NOW(), i.created_on) as age_days,
                COALESCE(DATEDIFF(NOW(), i.easy_status_updated_on), DATEDIFF(NOW(), i.updated_on)) as stale_days,
                CONCAT_WS(' ', u.lastname, u.firstname) as assigned_to_name,
                DATE_FORMAT(i.created_on, '%%Y-%%m-%%d %%H:%%i') as created_on
            FROM redmine.issues i
            JOIN redmine.issue_statuses s ON s.id = i.status_id
            LEFT JOIN redmine.users u ON u.id = i.assigned_to_id
            {where_clause}
            ORDER BY stale_days DESC
        """

        with conn.cursor() as cursor:
            cursor.execute(sql, params)
            rows = cursor.fetchall()

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df = pd.DataFrame(rows)
            if not df.empty:
                df.columns = ['ID', 'Тема', 'Статус', 'Возраст (дн)', 'Без движения (дн)', 'Исполнитель', 'Создана']
                df.to_excel(writer, sheet_name='Зависшие задачи', index=False)
                worksheet = writer.sheets['Зависшие задачи']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            else:
                writer.book.create_sheet('Зависшие задачи')
                writer.sheets['Зависшие задачи'].cell(row=1, column=1, value="Нет зависших задач")

        output.seek(0)
        filename = f"stale_tasks_{project_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        logger.error("Stale tasks export Error: %s", e, exc_info=True)
        return str(e), 500
    finally:
        if 'conn' in locals() and conn:
            conn.close()
